from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from analyzer import analyze_xml
from openpyxl import Workbook
import tempfile

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def home():
    return FileResponse("static/index.html")

@app.post("/analyze")
async def analyze(files: list[UploadFile] = File(...)):
    total_usage = {}
    all_timeline = []
    all_overlaps = []
    all_unused = set()

    for f in files:
        content = await f.read()
        usage, timeline, overlaps, unused = analyze_xml(content)

        for k, v in usage.items():
            total_usage[k] = total_usage.get(k, 0) + v

        all_timeline.extend(timeline)
        all_overlaps.extend(overlaps)
        all_unused.update(unused)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["File", "Seconds"])
    for k,v in total_usage.items():
        ws1.append([k,v])

    ws2 = wb.create_sheet("Timeline")
    ws2.append(["File","Start","End","Duration"])
    for t in all_timeline:
        ws2.append([t["file"], t["start"], t["end"], t["duration"]])

    ws3 = wb.create_sheet("Overlaps")
    ws3.append(["File","Start","End"])
    for o in all_overlaps:
        ws3.append([o["file"], o["start"], o["end"]])

    ws4 = wb.create_sheet("Unused")
    for u in all_unused:
        ws4.append([u])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)

    return FileResponse(tmp.name, filename="analysis.xlsx")
